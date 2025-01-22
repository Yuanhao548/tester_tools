import random
from comfunc.mock_data import *
from openpyxl.reader.excel import load_workbook


def read_excel_by_row(file_name, row_index):
    # 读取 Excel 文件，这里假设文件名为 'example.xlsx'
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # 读取第 3 行数据（注意：行索引从 1 开始）
    row_data = []
    for cell in sheet[row_index]:
        row_data.append(cell.value)
    print('参数行：', row_data)
    return row_data

def load_excel_and_write_mock(file_name, data, start_row, total, custom_value_indexs):
    # 加载现有的工作簿并批量写入mock数据
    workbook = load_workbook(file_name)
    # 选择工作表
    sheet = workbook.active
    # 向单元格中填充数据
    for _t in range(total):
        for _i, _d in enumerate(data):
            if _i not in custom_value_indexs:
                sheet.cell(row=_t+start_row, column=_i+1, value=eval(_d))
            else:
                _d = _d.replace('，', ',')
                value = random.choice(_d.split(','))
                sheet.cell(row=_t+start_row, column=_i+1, value=value)
    # 保存工作簿
    workbook.save(file_name)